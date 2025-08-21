# sistema_pontuacao_flask.py

from flask import Flask, render_template, request, redirect, flash, url_for
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
import psycopg2
import os
import re
import cloudinary
import cloudinary.uploader
import tempfile
import zipfile
import pandas as pd
from flask import send_file
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
DELETE_PASSWORD = 'confie123'



app = Flask(__name__)
app.secret_key = 'confie'

# Conexão com o banco PostgreSQL no Render
def get_db_connection():
    return psycopg2.connect(os.environ['DATABASE_URL'])

# Inicializa o banco e cria as tabelas se não existirem
def init_db():
    conn = get_db_connection()
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS pontuacoes (
            id SERIAL PRIMARY KEY,
            data TEXT,
            setor TEXT,
            obrigacao TEXT,
            pontuacao TEXT,
            observacao TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS loja (
            id SERIAL PRIMARY KEY,
            data TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS expedicao (
            id SERIAL PRIMARY KEY,
            data TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS logistica (
            id SERIAL PRIMARY KEY,
            data TEXT,
            motorista TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS comercial (
            id SERIAL PRIMARY KEY,
            data TEXT,
            vendedor TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')

    conn.commit()
    conn.close()

def fazer_backup_e_enviar():
    try:
        conn = get_db_connection()
        c = conn.cursor()

        tabelas = ['loja', 'expedicao', 'logistica', 'comercial', 'pontuacoes']
        arquivos_csv = []

        with tempfile.TemporaryDirectory() as tmpdirname:
            for tabela in tabelas:
                c.execute(f"SELECT * FROM {tabela}")
                rows = c.fetchall()
                colnames = [desc[0] for desc in c.description]

                if not rows:
                    continue

                # Salva CSV temporário
                df = pd.DataFrame(rows, columns=colnames)
                caminho_csv = os.path.join(tmpdirname, f"{tabela}.csv")
                df.to_csv(caminho_csv, index=False)
                arquivos_csv.append(caminho_csv)

            # Compactar todos os arquivos em um ZIP
            caminho_zip = os.path.join(tmpdirname, f"backup_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.zip")
            with zipfile.ZipFile(caminho_zip, 'w') as zipf:
                for file in arquivos_csv:
                    zipf.write(file, os.path.basename(file))

            # Enviar para o Cloudinary
            resultado = cloudinary.uploader.upload(
                caminho_zip,
                resource_type='raw',
                folder='backups_pontuacao',
                use_filename=True,
                unique_filename=False,
                overwrite=False
            )

            return resultado['secure_url']

    except Exception as e:
        print("Erro ao fazer backup automático:", e)
        return None

# Conversor de data (dd/mm/aaaa ou yyyy-mm-dd → yyyy-mm-dd)
def norm_date_to_iso(s):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return None


@app.template_filter('datetimeformat')
def datetimeformat(value, format='%d/%m/%Y'):
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime(format)
    except:
        return value

@app.route('/')
def home():
    conn = get_db_connection()
    c = conn.cursor()

    setores = []

    # Loja (sem média)
    c.execute("SELECT total FROM loja")
    loja_pontos = [row[0] for row in c.fetchall()]
    setores.append({
        'nome': 'Loja',
        'total_registros': len(loja_pontos),
        'soma': sum(loja_pontos) if loja_pontos else 0,
        'media': None,
        'valor_grafico': sum(loja_pontos) if loja_pontos else 0
    })

    # Expedição (sem média)
    c.execute("SELECT total FROM expedicao")
    expedicao_pontos = [row[0] for row in c.fetchall()]
    setores.append({
        'nome': 'Expedição',
        'total_registros': len(expedicao_pontos),
        'soma': sum(expedicao_pontos) if expedicao_pontos else 0,
        'media': None,
        'valor_grafico': sum(expedicao_pontos) if expedicao_pontos else 0
    })

    # Logística (usa média)
    c.execute("SELECT total FROM logistica")
    logistica_pontos = [row[0] for row in c.fetchall()]
    soma_log = sum(logistica_pontos) if logistica_pontos else 0
    media_log = round(soma_log / 6, 2) if soma_log else 0
    setores.append({
        'nome': 'Logística',
        'total_registros': len(logistica_pontos),
        'soma': soma_log,
        'media': media_log,
        'valor_grafico': media_log
    })

    # Comercial (usa média)
    c.execute("SELECT total FROM comercial")
    comercial_pontos = [row[0] for row in c.fetchall()]
    soma_com = sum(comercial_pontos) if comercial_pontos else 0
    media_com = round(soma_com / 8, 2) if soma_com else 0
    setores.append({
        'nome': 'Comercial',
        'total_registros': len(comercial_pontos),
        'soma': soma_com,
        'media': media_com,
        'valor_grafico': media_com
    })

    c.close()
    conn.close()

    return render_template(
        'home.html',
        total_loja=setores[0]['total_registros'],
        soma_loja=setores[0]['soma'],
        total_expedicao=setores[1]['total_registros'],
        soma_expedicao=setores[1]['soma'],
        total_logistica=setores[2]['total_registros'],
        soma_logistica=setores[2]['soma'],
        media_logistica=setores[2]['media'],
        total_comercial=setores[3]['total_registros'],
        soma_comercial=setores[3]['soma'],
        media_comercial=setores[3]['media']
    )


@app.route('/enviar', methods=['POST'])
def enviar():
    setor = request.form['setor']
    obrigacao = request.form['obrigacao']
    pontuacao = request.form['pontuacao']
    observacao = request.form.get('observacao', '')
    data = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db_connection()
    c = conn.cursor()

    c.execute('INSERT INTO pontuacoes (data, setor, obrigacao, pontuacao, observacao) VALUES (%s, %s, %s, %s, %s)',
              (data, setor, obrigacao, pontuacao, observacao))
    conn.commit()
    conn.close()

    flash("✅ Pontuação registrada com sucesso!", "success")
    return redirect('/')

@app.route('/historico')
def historico():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('SELECT data, setor, obrigacao, pontuacao, observacao FROM pontuacoes ORDER BY data DESC')
    registros = c.fetchall()
    conn.close()
    return render_template('historico.html', registros=registros)

# =======================================================================
# LOJA
# =======================================================================
# Função para converter valores de forma segura
def safe_int(value):
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0

# Dentro da rota de envio (exemplo para Loja)
@app.route('/loja', methods=['GET', 'POST'])
def loja():
    if request.method == 'POST':
        data_unica = request.form.get('data', '').strip()
        criterios  = request.form.getlist('criterios')  # checkboxes A-E
        observacao = request.form.get('observacao', '')
        extras     = request.form.getlist('extras')

        # Pesos fixos de cada critério (igual ao seu)
        pesos = {'A': 1, 'B': 1, 'C': 1, 'D': -1, 'E': -2}

        # Flags (1 se marcou, 0 se não) – igual ao seu
        A = int('A' in criterios)
        B = int('B' in criterios)
        C = int('C' in criterios)
        D = int('D' in criterios)
        E = int('E' in criterios)

        # === NOVO: ler múltiplas datas (opcional) com suporte a dd/mm/aaaa ===
        datas_raw = request.form.get('datas', '').strip()
        lista_datas, invalidas = [], []

        if datas_raw:
            # aceita vírgula, espaço ou quebra de linha
            tokens = re.split(r'[,\n;\s]+', datas_raw)
            for t in tokens:
                if not t:
                    continue
                iso = norm_date_to_iso(t)  # aceita dd/mm/aaaa ou yyyy-mm-dd
                if iso:
                    lista_datas.append(iso)
                else:
                    invalidas.append(t)

        # Se não veio múltipla, usa a data única (também normalizada)
        if not lista_datas:
            iso = norm_date_to_iso(data_unica or '')
            if not iso:
                flash('❌ Informe a data ou selecione múltiplas datas no formato dd/mm/aaaa.', 'danger')
                return redirect('/loja')
            lista_datas = [iso]

        # Evita datas repetidas e ordena
        lista_datas = sorted(set(lista_datas))

        inseridos = 0
        pulados   = 0

        conn = get_db_connection()
        c = conn.cursor()
        try:
            for dia in lista_datas:
                # Mesmas travas que você já usa: A–E não podem repetir no mesmo dia
                c.execute("SELECT A, B, C, D, E FROM loja WHERE data = %s", (dia,))
                registros = c.fetchall()

                conflito = False
                for registro in registros:
                    crits_existentes = {
                        'A': registro[0],
                        'B': registro[1],
                        'C': registro[2],
                        'D': registro[3],
                        'E': registro[4],
                    }
                    for c_sel in criterios:
                        if crits_existentes.get(c_sel, 0) == 1:
                            conflito = True
                            break
                    if conflito:
                        break

                if conflito:
                    # Em vez de abortar tudo, só pula esse dia
                    pulados += 1
                    continue

                # Total por dia (mesmas regras + extras)
                total = sum([pesos[c] for c in criterios])
                if 'meta' in extras:
                    total += 2
                if 'equipe90' in extras:
                    total += 1

                # Inserir no banco para este dia
                c.execute("""
                    INSERT INTO loja (data, A, B, C, D, E, extras, observacao, total)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (dia, A, B, C, D, E, ','.join(extras), observacao, total))
                inseridos += 1

            conn.commit()
        finally:
            conn.close()

        # Resumo amigável
        msgs = []
        if inseridos:
            msgs.append(f"✅ {inseridos} registro(s) inserido(s).")
        if pulados:
            msgs.append(f"⚠️ {pulados} dia(s) pulado(s) por já conterem os mesmos critérios.")
        if invalidas:
            msgs.append(f"❌ Datas inválidas ignoradas: {', '.join(invalidas)}")

        flash(' '.join(msgs) if msgs else "Nada a fazer.", "success" if inseridos else "warning")

        # mantém seu backup automático
        fazer_backup_e_enviar()
        return redirect('/loja')

    return render_template('loja.html')


# =======================================================================
# EXPEDIÇÃO
# =======================================================================
@app.route('/expedicao', methods=['GET', 'POST'])
def expedicao():
    if request.method == 'POST':
        data_unica = request.form.get('data', '').strip()
        criterios = request.form.getlist('criterios')

        A = 1  if 'A' in criterios else 0
        B = 1  if 'B' in criterios else 0
        C = 1  if 'C' in criterios else 0
        D = -2 if 'D' in criterios else 0
        E = -1 if 'E' in criterios else 0

        observacao = request.form.get('observacao', '')
        extras     = request.form.getlist('extras')

        extras_pontos = 0
        if 'meta' in extras:
            extras_pontos += 2
        if 'equipe90' in extras:
            extras_pontos += 1

        total_base = A + B + C + D + E + extras_pontos

        # === NOVO: múltiplas datas (aceita dd/mm/aaaa) ===
        datas_raw = request.form.get('datas', '').strip()
        lista_datas, invalidas = [], []

        if datas_raw:
            import re
            tokens = re.split(r'[,\n;\s]+', datas_raw)  # vírgula, espaço ou quebra de linha
            for t in tokens:
                if not t:
                    continue
                iso = norm_date_to_iso(t)  # usa seu helper (dd/mm/aaaa ou yyyy-mm-dd -> yyyy-mm-dd)
                if iso:
                    lista_datas.append(iso)
                else:
                    invalidas.append(t)

        if not lista_datas:
            iso = norm_date_to_iso(data_unica or '')
            if not iso:
                flash('❌ Informe a data ou selecione múltiplas datas no formato dd/mm/aaaa.', 'danger')
                return redirect('/expedicao')
            lista_datas = [iso]

        # Evita datas duplicadas
        lista_datas = sorted(set(lista_datas))

        inseridos, pulados = 0, []

        conn = get_db_connection()
        c = conn.cursor()
        try:
            for dia in lista_datas:
                # ✅ Travas por dia (mesmo comportamento de antes)
                c.execute("SELECT A, B, C, D, E FROM expedicao WHERE data = %s", (dia,))
                registros_dia = c.fetchall()

                conflito = (
                    ('A' in criterios and any(r[0] == 1   for r in registros_dia)) or
                    ('B' in criterios and any(r[1] == 1   for r in registros_dia)) or
                    ('C' in criterios and any(r[2] == 1   for r in registros_dia)) or
                    ('D' in criterios and any(r[3] == -2  for r in registros_dia)) or
                    ('E' in criterios and any(r[4] == -1  for r in registros_dia))
                )

                if conflito:
                    pulados.append(dia)
                    continue

                # INSERT mantendo sua estrutura
                c.execute("""
                    INSERT INTO expedicao (data, A, B, C, D, E, extras, observacao, total)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (dia, A, B, C, D, E, ','.join(extras), observacao, total_base))
                inseridos += 1

            conn.commit()
        finally:
            conn.close()

        # Feedback
        msgs = []
        if inseridos:
            msgs.append(f"✅ {inseridos} registro(s) inserido(s).")
        if pulados:
            msgs.append(f"⚠️ Dias pulados por já conterem os mesmos critérios: {', '.join(pulados)}.")
        if invalidas:
            msgs.append(f"❌ Datas inválidas ignoradas: {', '.join(invalidas)}")

        flash(' '.join(msgs) if msgs else "Nada a fazer.", "success" if inseridos else "warning")
        fazer_backup_e_enviar()
        return redirect('/expedicao')

    return render_template('expedicao.html')

@app.route('/historico_expedicao')
def historico_expedicao():
    conn = get_db_connection()
    c = conn.cursor()

    c.execute("SELECT id, data, A, B, C, D, E, extras, total, observacao FROM expedicao ORDER BY data DESC")
    registros = c.fetchall()

    total_geral = sum([r[8] for r in registros])  # índice do total
    conn.close()

    return render_template('historico_expedicao.html', registros=registros, total_geral=total_geral)
	

# Nova rota para a Logística com menu de motoristas e formulário de pontuação
@app.route('/logistica', methods=['GET', 'POST'])
def logistica():
    conn = get_db_connection()
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS logistica (
            id SERIAL PRIMARY KEY,
            data TEXT,
            motorista TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')

    motoristas = ['Denilson', 'Fabio', 'Rogerio', 'Robson', 'Simone', 'Vinicius', 'Equipe']

    if request.method == 'POST':
        data = request.form['data']
        motorista = request.form['motorista']
        A = safe_int(request.form.get('A'))
        B = safe_int(request.form.get('B'))
        C = safe_int(request.form.get('C'))
        D = safe_int(request.form.get('D'))
        E = safe_int(request.form.get('E'))
        extras = request.form.getlist('extras')
        observacao = request.form.get('observacao', '')

        total = A + B + C + D + E

        # ✅ Soma extra de economia
        if 'economia' in extras:
            total += 2

        # 🔒 Validação do extra 'equipe90'
        if 'equipe90' in extras and motorista != 'Equipe':
            flash("❌ O ponto extra 'Equipe chegou a 90%' só pode ser usado com o motorista 'Equipe'.", "danger")
            conn.close()
            return redirect('/logistica')

        if 'equipe90' in extras:
            total += 1

        # 🔒 Verificações de trava por motorista
        c.execute("SELECT A, B, C, D, E FROM logistica WHERE data = %s AND motorista = %s", (data, motorista))
        registros = c.fetchall()

        if any(r[0] == 1 for r in registros) and A == 1:
            flash("⚠️ O critério A já foi registrado para esse motorista nesse dia.", "danger")
            conn.close()
            return redirect('/logistica')
        if any(r[1] == 1 for r in registros) and B == 1:
            flash("⚠️ O critério B já foi registrado para esse motorista nesse dia.", "danger")
            conn.close()
            return redirect('/logistica')
        if any(r[2] == 1 for r in registros) and C == 1:
            flash("⚠️ O critério C já foi registrado para esse motorista nesse dia.", "danger")
            conn.close()
            return redirect('/logistica')
        if any(r[3] == -2 for r in registros) and D == -2:
            flash("⚠️ O critério D já foi registrado para esse motorista nesse dia.", "danger")
            conn.close()
            return redirect('/logistica')
        if any(r[4] == -1 for r in registros) and E == -1:
            flash("⚠️ O critério E já foi registrado para esse motorista nesse dia.", "danger")
            conn.close()
            return redirect('/logistica')

        # ✅ Se passou nas travas, insere
        c.execute('''
            INSERT INTO logistica (data, motorista, A, B, C, D, E, extras, observacao, total)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (data, motorista, A, B, C, D, E, ','.join(extras), observacao, total))

        conn.commit()
        conn.close()
        flash("✅ Pontuação da logística registrada com sucesso!", "success")
        fazer_backup_e_enviar()
        return redirect('/logistica')

    conn.close()
    return render_template('logistica.html', motoristas=motoristas)


@app.route('/historico_logistica')
def historico_logistica():
    conn = get_db_connection()
    c = conn.cursor()

    motorista = request.args.get('motorista', '')
    motoristas = ['Denilson', 'Fabio', 'Rogerio', 'Robson', 'Simone', 'Vinicius', 'Equipe']

    if motorista:
        c.execute('''
            SELECT id, data, motorista, A, B, C, D, E, extras, observacao, total  
            FROM logistica 
            WHERE motorista = %s 
            ORDER BY data DESC
        ''', (motorista,))
    else:
        c.execute('''
            SELECT id, data, motorista, A, B, C, D, E, extras, observacao, total 
            FROM logistica 
            ORDER BY data DESC
        ''')

    registros = c.fetchall()
    conn.close()

    total_geral = sum([int(r[10]) for r in registros]) if registros else 0

    if motorista:
        media = Decimal(total_geral).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    else:
        motoristas_reais = [m for m in motoristas if m != 'Equipe']
        media = (
            Decimal(total_geral / len(motoristas_reais)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            if motoristas_reais else Decimal('0.00')
        )

    return render_template(
        'historico_logistica.html',
        registros=registros,
        motorista=motorista,
        motoristas=motoristas,
        total_geral=total_geral,
        media=media
    )


@app.route('/historico_loja')
def historico_loja():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT id, data, A, B, C, D, E, extras, total, observacao FROM loja ORDER BY data DESC")
    registros = c.fetchall()
    conn.close()

    total_geral = sum([r[8] for r in registros]) if registros else 0  # índice 8 = campo 'total'
    media = round(total_geral / len(registros), 1) if registros else 0

    return render_template('historico_loja.html', registros=registros, total_geral=total_geral, media=media)

# =======================================================================
# COMERCIAL
# =======================================================================
@app.route('/comercial', methods=['GET', 'POST'])
def comercial():
    conn = get_db_connection()
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS comercial (
            id SERIAL PRIMARY KEY,
            data TEXT,
            vendedor TEXT,
            A INTEGER,
            B INTEGER,
            C INTEGER,
            D INTEGER,
            E INTEGER,
            extras TEXT,
            observacao TEXT,
            total INTEGER
        )
    ''')

    vendedores = ['EVERTON', 'MARCELO', 'PEDRO', 'SILVANA', 'TIAGO', 'RODOLFO', 'MARCOS', 'THYAGO', 'EQUIPE']

    if request.method == 'POST':
        data = request.form['data']
        vendedor = request.form['vendedor']
        A = safe_int(request.form.get('A'))
        B = safe_int(request.form.get('B'))
        C = safe_int(request.form.get('C'))
        D = safe_int(request.form.get('D'))
        E = safe_int(request.form.get('E'))
        extras = request.form.getlist('extras')
        observacao = request.form.get('observacao', '')

        # 🔒 Validação do ponto extra equipe90
        if 'equipe90' in extras and vendedor.upper() != 'EQUIPE':
            flash("❌ O ponto extra 'Equipe chegou a 90%' só pode ser usado com o vendedor 'EQUIPE'.", "danger")
            conn.close()
            return redirect('/comercial')

        # 🔒 Travas por vendedor e data (agora considerando qualquer valor diferente de zero)
        c.execute("SELECT A, B, C, D, E FROM comercial WHERE data = %s AND vendedor = %s", (data, vendedor))
        registros = c.fetchall()

        if A != 0 and any(r[0] != 0 for r in registros):
            flash("⚠️ O critério A já foi registrado para esse vendedor nesse dia.", "danger")
            conn.close()
            return redirect('/comercial')
        if B != 0 and any(r[1] != 0 for r in registros):
            flash("⚠️ O critério B já foi registrado para esse vendedor nesse dia.", "danger")
            conn.close()
            return redirect('/comercial')
        if C != 0 and any(r[2] != 0 for r in registros):
            flash("⚠️ O critério C já foi registrado para esse vendedor nesse dia.", "danger")
            conn.close()
            return redirect('/comercial')
        if D != 0 and any(r[3] != 0 for r in registros):
            flash("⚠️ O critério D já foi registrado para esse vendedor nesse dia.", "danger")
            conn.close()
            return redirect('/comercial')
        if E != 0 and any(r[4] != 0 for r in registros):
            flash("⚠️ O critério E já foi registrado para esse vendedor nesse dia.", "danger")
            conn.close()
            return redirect('/comercial')

        total = A + B + C + D + E
        if 'meta' in extras:
            total += 2
        if 'equipe90' in extras:
            total += 1

        c.execute('''
            INSERT INTO comercial (data, vendedor, A, B, C, D, E, extras, observacao, total)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (data, vendedor, A, B, C, D, E, ','.join(extras), observacao, total))

        conn.commit()
        conn.close()

        flash("✅ Pontuação registrada com sucesso!", "success")
        fazer_backup_e_enviar()
        return redirect('/comercial')

    conn.close()
    return render_template('comercial.html', vendedores=vendedores)



@app.route('/historico_comercial')
def historico_comercial():
    conn = get_db_connection()
    c = conn.cursor()

    vendedor = request.args.get('vendedor', '')

    # Lista fixa de vendedores
    lista_vendedores = ['EVERTON', 'MARCELO', 'PEDRO', 'SILVANA', 'TIAGO', 'RODOLFO', 'MARCOS', 'THYAGO', 'EQUIPE']

    # Busca registros com ou sem filtro
    if vendedor:
        c.execute("SELECT * FROM comercial WHERE vendedor = %s ORDER BY data DESC", (vendedor,))
    else:
        c.execute("SELECT * FROM comercial ORDER BY data DESC")

    registros = c.fetchall()

    # Cálculo do total geral e média (com base no número de registros)
    total_geral = sum([r[10] for r in registros])  # campo total = índice 10
    media = Decimal(total_geral / 8).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    conn.close()

    return render_template(
        'historico_comercial.html',
        registros=registros,
        total_geral=total_geral,
        media=media,
        vendedores=lista_vendedores,
        vendedor=vendedor
    )


@app.route('/zerar_tudo', methods=['POST'])
def zerar_tudo():
    senha = request.form.get('senha')
    if senha == "confie123":  # ajuste para sua senha desejada
        conn = get_db_connection()
        c = conn.cursor()
        for tabela in ['loja', 'expedicao', 'logistica', 'comercial']:
            c.execute(f"DELETE FROM {tabela}")
        conn.commit()
        conn.close()
        flash("✅ Todas as pontuações foram zeradas com sucesso!", "success")
    else:
        flash("❌ Senha incorreta. Ação cancelada.", "danger")

    return redirect('/')


@app.route('/criar_banco')
def criar_banco():
    try:
        init_db()
        return "✅ Banco de dados criado com sucesso!"
    except Exception as e:
        return f"❌ Erro ao criar banco: {str(e)}"

@app.route('/restaurar_backup', methods=['GET', 'POST'])
def restaurar_backup():
    if request.method == 'POST':
        arquivo = request.files['backup']
        if arquivo and arquivo.filename.endswith('.zip'):
            try:
                with tempfile.TemporaryDirectory() as tmpdirname:
                    caminho_zip = os.path.join(tmpdirname, arquivo.filename)
                    arquivo.save(caminho_zip)

                    # Extrair ZIP
                    with zipfile.ZipFile(caminho_zip, 'r') as zip_ref:
                        zip_ref.extractall(tmpdirname)

                    # Conectar ao banco
                    conn = get_db_connection()
                    c = conn.cursor()

                    tabelas = ['loja', 'expedicao', 'logistica', 'comercial', 'pontuacoes']
                    for tabela in tabelas:
                        caminho_csv = os.path.join(tmpdirname, f"{tabela}.csv")
                        if os.path.exists(caminho_csv):
                            df = pd.read_csv(caminho_csv)

                            # Limpa a tabela
                            c.execute(f"DELETE FROM {tabela}")

                            # Insere os dados
                            for _, row in df.iterrows():
                                colunas = ','.join(df.columns)
                                valores = ','.join(['%s'] * len(df.columns))
                                c.execute(f"INSERT INTO {tabela} ({colunas}) VALUES ({valores})", tuple(row))

                    conn.commit()
                    conn.close()
                    flash("✅ Backup restaurado com sucesso!", "success")
                    return redirect('/')
            except Exception as e:
                flash(f"❌ Erro ao restaurar backup: {e}", "danger")
                return redirect('/restaurar_backup')

    return render_template('restaurar_backup.html')

@app.route('/baixar_relatorio_excel')
def baixar_relatorio_excel():
    conn = get_db_connection()
    c = conn.cursor()

    tabelas = ['loja', 'expedicao', 'logistica', 'comercial', 'pontuacoes']
    nomes_formatados = {
        'loja': {
            'A': 'Organização da loja (Gerente ADM)',
            'B': 'Pontualidade (RH)',
            'C': 'Fechamento do caixa (Financeiro)',
            'D': 'Não postar em rede social (RH)',
            'E': 'Validade / Avaria (Gerente ADM)'
        },
        'expedicao': {
            'A': 'Organização estoque (Gerente ADM)',
            'B': 'Separação correta (Faturamento)',
            'C': 'Faturamento OK (Financeiro)',
            'D': 'Erros / Devoluções (Financeiro)',
            'E': 'Finalização após horário'
        },
        'logistica': {
            'A': 'Separação Correta',
            'B': 'Entrega Realizada',
            'C': 'Roteiro Otimizado',
            'D': 'Veículo limpo/organizado',
            'E': 'Sem reclamações'
        },
        'comercial': {
            'A': 'Acompanhamento pedidos',
            'B': 'Prospecção ativa',
            'C': 'Metas diárias',
            'D': 'Ajustes manuais',
            'E': 'Participação reuniões'
        }
    }

    wb = Workbook()
    wb.remove(wb.active)

    for tabela in tabelas:
        c.execute(f"SELECT * FROM {tabela}")
        dados = c.fetchall()
        colunas = [desc[0] for desc in c.description]

        if not dados:
            continue

        ws = wb.create_sheet(title=tabela.capitalize())

        # Cabeçalhos amigáveis
        headers = []
        for col in colunas:
            if tabela in nomes_formatados and col.upper() in nomes_formatados[tabela]:
                headers.append(nomes_formatados[tabela][col.upper()])
            else:
                headers.append(col.capitalize())
        ws.append(headers)

        # Estilo do cabeçalho
        for col in ws[1]:
            col.font = Font(bold=True, color="FFFFFF")
            col.fill = PatternFill(start_color="1f4e78", end_color="1f4e78", fill_type="solid")
            col.alignment = Alignment(horizontal="center", vertical="center")

        total_geral = 0
        for linha in dados:
            linha_formatada = []
            for i, cell in enumerate(linha):
                if cell is None or str(cell).lower() == 'nan':
                    linha_formatada.append('')
                elif isinstance(cell, str) and cell.lower() == 'nan':
                    linha_formatada.append('')
                elif isinstance(cell, datetime):
                    linha_formatada.append(cell.strftime("%d/%m/%Y"))
                else:
                    linha_formatada.append(cell)
            ws.append(linha_formatada)

            if 'total' in colunas:
                total_geral += linha[colunas.index('total')]

        # Linha do total
        if 'total' in colunas:
            idx_total = colunas.index('total') + 1
            ws.append([""] * (idx_total - 1) + ["Total Geral:", total_geral])
            ultima_linha = ws.max_row
            ws[f"{chr(64 + idx_total)}{ultima_linha}"].font = Font(bold=True, color="1f4e78")

        # Estilo geral
        thin = Side(border_style="thin", color="999999")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Ajuste de largura
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    conn.close()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'relatorio_pontuacoes_{datetime.now().strftime("%Y-%m-%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/deletar', methods=['GET', 'POST'])
def deletar():
    if request.method == 'POST':
        tabela = request.form.get('tabela')
        id_registro = request.form.get('id')
        senha = request.form.get('senha')

        if senha != DELETE_PASSWORD:
            flash("❌ Senha incorreta.", "danger")
            return redirect('/deletar')

        if tabela not in ['loja', 'expedicao', 'logistica', 'comercial']:
            flash("❌ Tabela inválida.", "danger")
            return redirect('/deletar')

        try:
            conn = get_db_connection()
            c = conn.cursor()
            c.execute(f"DELETE FROM {tabela} WHERE id = %s", (id_registro,))
            conn.commit()
            conn.close()
            flash(f"✅ Registro ID {id_registro} apagado da tabela {tabela}.", "success")
        except Exception as e:
            flash(f"❌ Erro ao deletar: {str(e)}", "danger")

        return redirect('/deletar')

    return render_template('deletar.html')


if __name__ == '__main__':
        app.run(debug=True)